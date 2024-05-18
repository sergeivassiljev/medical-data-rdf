import openpyxl
import re

source_file_path = 'request_responses_6.xlsx'
wb = openpyxl.load_workbook(source_file_path)
ws = wb.active

new_wb = openpyxl.Workbook()
new_ws = new_wb.active

def clean_text(text):
    return re.sub(r'[^\w\s]', '', text)  

for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=1), start=2):
    for cell in row:
        cleaned_text = clean_text(cell.value) if cell.value else ''
        new_ws[f'A{idx}'].value = cleaned_text
        
if new_ws.max_row >= 1:
    new_ws.delete_rows(1)
    
new_file_path = 'cleared_text_6.xlsx'
new_wb.save(new_file_path)
wb.save('')

